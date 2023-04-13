import tkinter.messagebox
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

background='#06283D'
framebg='#EDEDED'
framefg='#06283D'

root=Tk()
root.title('Student Registration From')
root.geometry('1200x610+20+10')
root.resizable(False,False)
root.config(bg=background)


file=pathlib.Path('student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] ="Registration No"
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Date of Birth"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religious"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Father's Occupation"
    sheet['K1'] = "Mother's Name"
    sheet['L1'] = "Mother's Occupation"
    sheet['M1'] = "Day"
    sheet['N1'] = "Month"
    sheet['O1'] = "Year"
    file.save('student_data.xlsx')


## show image
def Showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title='Select image file',filetypes=(('JPG File','*.jpg'),
                                                                             ('PNG File','*.png'),
                                                                             ('All File','*.txt')))
    img=(Image.open(filename))
    resize_img=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resize_img)
    ibl.config(image=photo2)
    ibl.image=photo2



############# registration##########
def registration_no():
    file=openpyxl.load_workbook('student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set('1')
#########clear####
def Clear():
    global img

    Name.set('')
    Day.set('Day')
    Month.set('Month')
    Year.set('Year')
    Class.set('Select Class')
    skills.set('Select Skill')
    riligion.set('Select Religion')
    FatherName.set('')
    Occupation.set('')
    MotherName.set('')
    MotherOccupation.set('')
    registration_no()
    save_button.config(state='normal')
    img1=PhotoImage(file='icon/upload photo.png')
    ibl.config(image=img1)
    ibl.image=img1
    img=""


#gender selection
def selection():
    global gender

    valu=radio.get()
    if valu==1:
        gender='Male'

    else:
        gender='Female'

######## save button ########

def Save():
    global gender

    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender

    except:
        messagebox.showerror('Error','Select Gender ')

    day=Day.get()
    month=Month.get()
    year=Year.get()
    D2=Date.get()

    religon=riligion.get()
    S1=skills.get()
    f_name=FatherName.get()
    foqu=Occupation.get()
    m_name=MotherName.get()
    moqu=MotherOccupation.get()
    if N1=='' or day=='Day' or month=='Month' or year=='Year' or religon=='Select Religion' or C1=='Select Class' or S1=='Select Skill' or f_name=='' or foqu=='' or moqu=='' or m_name=='':
        messagebox.showerror('Error','Some data are missing')
    else:
        file=openpyxl.load_workbook('student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2, row=sheet.max_row , value=N1)
        sheet.cell(column=3, row=sheet.max_row , value=C1)
        sheet.cell(column=4, row=sheet.max_row , value=gender)
        sheet.cell(column=5, row=sheet.max_row , value=str(day+' '+month+' '+year))
        sheet.cell(column=6, row=sheet.max_row , value=D2)
        sheet.cell(column=7, row=sheet.max_row , value=religon)
        sheet.cell(column=8, row=sheet.max_row , value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=f_name)
        sheet.cell(column=10, row=sheet.max_row , value=foqu)
        sheet.cell(column=11, row=sheet.max_row , value=m_name)
        sheet.cell(column=12, row=sheet.max_row , value=moqu)
        sheet.cell(column=13,row=sheet.max_row,value=day)
        sheet.cell(column=14, row=sheet.max_row, value=month)
        sheet.cell(column=15, row=sheet.max_row, value=year)


        file.save(r'student_data.xlsx')
        try:
            img.save('student image/'+str(R1)+'.jpg')
        except:
            messagebox.showinfo('info','profile picture is not Avilable')
        messagebox.showinfo('info','successfully data entred')
        Clear()
        registration_no()

############# update button program ################
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()

    selection()
    G1=gender

    day = Day.get()
    month = Month.get()
    year = Year.get()
    D2 = Date.get()

    religon = riligion.get()
    S1 = skills.get()
    f_name = FatherName.get()
    foqu = Occupation.get()
    m_name = MotherName.get()
    moqu = MotherOccupation.get()
    file=openpyxl.load_workbook('student_data.xlsx')
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_no=str(name)[15:-1]
    # sheet.cell( column=1,row=int(reg_no),value=R1)
    try:
        sheet.cell(column=2,row=int(reg_no), value=N1)
        sheet.cell(column=3,row=int(reg_no),value=C1)
        sheet.cell(column=4,row=int(reg_no), value=G1)
        sheet.cell(column=5,row=int(reg_no), value=str(day+' '+month+' '+year))
        sheet.cell(column=6,row=int(reg_no), value=D2)
        sheet.cell(column=7,row=int(reg_no), value=religon)
        sheet.cell(column=8,row=int(reg_no), value=S1)
        sheet.cell(column=9,row=int(reg_no), value=f_name)
        sheet.cell(column=10,row=int(reg_no), value=foqu)
        sheet.cell(column=11,row=int(reg_no), value=m_name)
        sheet.cell(column=12,row=int(reg_no), value=moqu)
        sheet.cell(column=13,row=int(reg_no), value=day)
        sheet.cell(column=14,row=int(reg_no), value=month)
        sheet.cell(column=15,row=int(reg_no), value=year)
        if N1 == '' or day == 'Day' or month == 'Month' or year == 'Year' or religon == 'Select Religion' or C1 == 'Select Class' or S1 == 'Select Skill' or f_name == '' or foqu == '' or moqu == '' or m_name == '':
            messagebox.showerror('Error','Some data are missing')
        else:
            file.save(r'student_data.xlsx')
            messagebox.showinfo('Update', 'Updated Successfully')
    except:
        messagebox.showerror('Error','None updated')
    try:
        img.save('student image/'+str(R1)+'.jpg')
    except:
        messagebox.showerror('Error','Entered correct one')

    Clear()


def search():

    try:
        text=Search.get()
        Clear()
        save_button.config(state='disable')
        file=openpyxl.load_workbook('student_data.xlsx')
        sheet=file.active
        for row in sheet.rows:
            if row[0].value==int(text):
                name=row[0]

                reg_no_position=str(name)[14:-1]

                reg_no=str(name)[15:-1]


        try:
            str(name)
        except:
            messagebox.showerror('Invalid','Invalid registration')

        x1=sheet.cell(row=int(reg_no),column=1).value
        x2 = sheet.cell(row=int(reg_no), column=2).value
        x3 = sheet.cell(row=int(reg_no), column=3).value
        x4 = sheet.cell(row=int(reg_no), column=4).value
        x5 = sheet.cell(row=int(reg_no), column=5).value
        x6 = sheet.cell(row=int(reg_no), column=6).value
        x7 = sheet.cell(row=int(reg_no), column=7).value
        x8 = sheet.cell(row=int(reg_no), column=8).value
        x9 = sheet.cell(row=int(reg_no), column=9).value
        x10 = sheet.cell(row=int(reg_no), column=10).value
        x11 = sheet.cell(row=int(reg_no), column=11).value
        x12 = sheet.cell(row=int(reg_no), column=12).value
        x13 = sheet.cell(row=int(reg_no), column=13).value
        x14 = sheet.cell(row=int(reg_no), column=14).value
        x15 = sheet.cell(row=int(reg_no), column=15).value



        Registration.set(x1)
        Name.set(x2)
        Class.set(x3)
        Day.set(x13)
        Month.set(x14)
        Year.set((x15))
        if x4=='Female':
            g2.select()
        else:
            g1.select()
        Date.set(x6)
        riligion.set(x7)
        skills.set(x8)
        FatherName.set(x9)
        Occupation.set(x10)
        MotherName.set(x11)
        MotherOccupation.set(x12)
        img=Image.open('student image/'+str(x1)+'.jpg')
        resize_image=img.resize((190,190))
        photo=ImageTk.PhotoImage(resize_image)
        ibl.config(image=photo)
        ibl.img=photo
    except:
        messagebox.showerror('Error','please Enter Registration no first')











### exit window
def Exit():
    ans=messagebox.askyesno(title='Exit',message='Do You Want To Exit ?')
    if ans:
        root.destroy()

root.protocol("WM_DELETE_WINDOW",Exit)


#top lebel
heading=Label(root,text='HetalKhali Nasirabad Darus Sunnah Dakhil Madrasha',bg='#f0687c',width=10,
      height=3,anchor=CENTER,font='Arial 15')
heading.pack(side=TOP,fill=X)


Label(root,text='Student Registration',bg='#c36464',fg='#fff',width=10,
      height=2,font='Arial 20 bold').pack(side=TOP,fill=X)
#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,border=2,font='arial 20').place(x=820,y=85)


imageicon3=PhotoImage(file="icon/icons8-google-web-search-24.png")
srch=Button(root,text='Search',image=imageicon3,compound=RIGHT,width=123,bg='#68ddfa',font='arial 13 bold',cursor='hand2',command=search)
srch.place(x=1060,y=88)

imageicon4=PhotoImage(file='icon/icons8-update-48.png')
update_button=Button(root,image=imageicon4,text='Update',compound=LEFT,border=1,cursor='hand2',command=Update)
update_button.place(x=64,y=85)
#Registration and date
Label(root,text='Registration No: ',fg=framebg,bg=background,font='arial 13').place(x=30,y=150)
Label(root,text='Date: ',fg=framebg,bg=background,font='arial 13').place(x=500,y=150)

Registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Registration,font='arial 10',width=15)
reg_entry.place(x=160,y=150)

registration_no()


#set date
today=date.today()

d1=today.strftime('%d/%m/%y')
date_entry=Entry(root,textvariable=Date,width=15,font='arial 10')
date_entry.place(x=550,y=150)
Date.set(d1)
date_entry.config(state='disable')

#student details
obj=LabelFrame(root,text="Student's Details",font=20,fg=framefg,bg=framebg,height=180,width=900,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text='Full Name :',font='arial 13',fg=framefg,bg=framebg).place(x=30,y=30)
Label(obj,text='Date of Birth :',font='arial 13',fg=framefg,bg=framebg).place(x=30,y=70)
Label(obj,text='Gender :',font='arial 13',fg=framefg,bg=framebg).place(x=30,y=110)

Label(obj,text='Class :',font='arial 13',fg=framefg,bg=framebg).place(x=500,y=30)
Label(obj,text='Religion :',font='arial 13',fg=framefg,bg=framebg).place(x=500,y=70)
Label(obj,text='Skills :',font='arial 13',fg=framefg,bg=framebg).place(x=500,y=110)

obj2=LabelFrame(root,text="Parent's Details",font=20,fg=framefg,bg=framebg,height=180,width=900,relief=GROOVE)
obj2.place(x=30,y=410)
Label(obj2,text="Father's Name :",font='arial 13',fg=framefg,bg=framebg).place(x=30,y=30)
Label(obj2,text="Occupation :",font='arial 13',fg=framefg,bg=framebg).place(x=30,y=70)
# Label(obj2,text="Father's Age :",font='arial 13',fg=framefg,bg=framebg).place(x=30,y=110)

FatherName=StringVar()
fathername_entry=Entry(obj2,textvariable=FatherName,width=20,font='arial 10')
fathername_entry.place(x=170,y=30)


Occupation=StringVar()
fatheroccupation_entry=Entry(obj2,textvariable=Occupation,width=20,font='arial 10')
fatheroccupation_entry.place(x=170,y=70)






Label(obj2,text="Mother's Name :",font='arial 13',fg=framefg,bg=framebg).place(x=500,y=30)
Label(obj2,text="Occupation :",font='arial 13',fg=framefg,bg=framebg).place(x=500,y=70)
# Label(obj2,text="Mother's Age :",font='arial 13',fg=framefg,bg=framebg).place(x=500,y=110)

MotherName=StringVar()
MotherName_entry=Entry(obj2,textvariable=MotherName,width=20,font='arial 10')
MotherName_entry.place(x=650,y=30)

MotherOccupation=StringVar()
motheroccupation_entry=Entry(obj2,textvariable=MotherOccupation,width=20,font='arial 10')
motheroccupation_entry.place(x=650,y=70)

####gender button
radio=IntVar()

g1=Radiobutton(obj,text='Male',variable=radio,fg=framefg,bg=framebg,value=1,cursor='hand2',command=lambda :selection())
g1.place(x=165,y=110)

g2=Radiobutton(obj,text='Female',variable=radio,fg=framefg,bg=framebg,value=2,cursor='hand2',command=lambda :selection())
g2.place(x=245,y=110)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font='arial 10')
name_entry.place(x=170,y=33)

# Dob=StringVar()
# dob_entry=Entry(obj,textvariable=Dob,width=20,font='arial 10')
# dob_entry.place(x=170,y=73)

Day=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25',
                '26','27','28','29','30','31'],font='Roboto 10 ',width=4,state='r',cursor='hand2')
Day.set('Day')
Day.place(x=170,y=73)

Month=Combobox(obj,values=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'],font='Roboto 10 ',width=5,state='r',cursor='hand2')
Month.set('Month')
Month.place(x=230,y=73)

Year=Combobox(obj,values=['2023', '2022', '2021', '2020', '2019', '2018', '2017', '2016', '2015', '2014', '2013', '2012', '2011', '2010', '2009', '2008', '2007', '2006', '2005', '2004', '2003', '2002', '2001', '2000', '1999', '1998', '1997', '1996', '1995', '1994', '1993', '1992', '1991', '1990'],font='Roboto 10 ',width=5,state='r',cursor='hand2')
Year.place(x=300,y=73)
Year.set('Year')

riligion=Combobox(obj,values=['Islam','Hindu','Others'],font='Roboto 10 ',width=17,state='r',cursor='hand2')
riligion.place(x=650,y=70)
riligion.set('Select Religion')


skills=Combobox(obj,values=['Cricket','Football','Badminton','Swimming'],font='Roboto 10 ',width=17,state='r',cursor='hand2')
skills.place(x=650,y=110)
skills.set('Select Skill')

Class=Combobox(obj,values=['Six','Seven','Eight','Nine','Ten'],font='Roboto 10 ',width=17,state='r',cursor='hand2')
Class.place(x=650,y=30)
Class.set('Select Class')

###################
f=Frame(root,bd=3,bg='black',width=198,height=180,relief=GROOVE)
f.place(x=990,y=150)
img=PhotoImage(file='icon/upload photo.png')
ibl=Label(f,image=img,bg='black')
ibl.place(x=0,y=0)

upload_button=Button(root,text='Upload',width=17,height=2,font='Arial 12 bold'
                     ,bg='lightblue',cursor='hand2',command=lambda :Showimage())
upload_button.place(x=1000,y=340)
save_button=Button(root,text='Save',width=17,height=2,font='Arial 12 bold',
                   bg='lightgreen',cursor='hand2',command=Save)
save_button.place(x=1000,y=405)
reset_button=Button(root,text='Reset',width=17,height=2,font='Arial 12 bold'
                    ,bg='lightpink',cursor='hand2',command=Clear)
reset_button.place(x=1000,y=475)
exit_button=Button(root,text='Exit',width=17,height=2,font='Arial 12 bold',
                   bg='lightgrey',cursor='hand2',command=lambda :Exit())
exit_button.place(x=1000,y=545)






root.mainloop()